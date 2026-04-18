"""
Research-Quality Evaluation Script
====================================
Compares ground_truth.csv vs master.xlsx for cricket bowling biomechanics.

Outputs
-------
1. evaluation_results.xlsx  – structured multi-sheet Excel workbook
2. scatter_plots.png        – identity-line scatter plots
3. bland_altman_plots.png   – Bland-Altman agreement plots
4. error_distribution.png   – error histograms + boxplots
5. summary_heatmap.png      – MAE / nRMSE heatmap across all parameters

Design notes
------------
* MAE / RMSE are reported in native units (always interpretable).
* nRMSE is normalised by the ground-truth range (max-min), giving a
  dimensionless [0-1] measure that is fair for bounded-range data such as
  angles (0-180°). This avoids the inflation you get with percentage-error
  when the GT value is small (e.g. 18°).
* Circular / arc statistics are NOT used here because the angles are elbow /
  knee joint angles stored as scalars in [0, 180], not wrap-around headings.
  Range-normalised RMSE is the appropriate alternative.
* Bland-Altman plots are the gold standard for method-comparison studies and
  are required by most sports-science / biomechanics journals.
* Intraclass Correlation Coefficient (ICC type 2,1) is also computed as a
  relative reliability index alongside the absolute error metrics.
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import openpyxl.utils.cell as opu

warnings.filterwarnings("ignore")

# ── Parameter catalogue ──────────────────────────────────────────────────────
# Categories and labels are aligned with the research proposal scope (Section 6):
#   Elbow Kinematics | Knee Kinematics | Step & Stride Timing |
#   Delivery Stride | Head COM | Wrist Velocity | Event Frames

ELBOW_PARAMS = [
    ("elbow_angle_arm_back_deg",  "Elbow Angle at Arm-Back",              "deg",   0, 180),
    ("elbow_angle_release_deg",   "Elbow Angle at Release",               "deg",   0, 180),
    ("elbow_extension_deg",       "Elbow Extension (Arm-Back to Release)","deg", -45,  45),
]

KNEE_PARAMS = [
    ("knee_angle_ffc_deg",        "Front Knee Angle at FFC",              "deg",   0, 180),
    ("knee_angle_release_deg",    "Front Knee Angle at Release",          "deg",   0, 180),
]

TEMPORAL_PARAMS = [
    ("stride_duration_s",         "Delivery Stride Duration",             "s",    0,   1),
    ("step_duration_mean_s",      "Step Duration (Mean)",                 "s",    0,   1),
    ("step_duration_std_s",       "Step Duration (SD)",                   "s",    0,   0.5),
    ("step_duration_cv",          "Step Duration (CV)",                   "",     0,   1),
    ("final5_total_duration_s",   "Final 5 Steps Total Duration",         "s",    0,   3),
]

STRIDE_PARAMS = [
    ("stride_length_m",           "Delivery Stride Length",               "m",    0,   2.5),
]

HEAD_COM_PARAMS = [
    ("head_dx_ffc_cm",            "Head COM Lateral Offset at FFC",       "cm", -80,  80),
    ("head_dy_ffc_cm",            "Head COM Vertical Offset at FFC",      "cm", -200,  0),
    ("head_d_ffc_cm",             "Head COM Distance at FFC",             "cm",  50, 200),
    ("head_dx_bfc_cm",            "Head COM Lateral Offset at BFC",       "cm", -100, 20),
    ("head_dy_bfc_cm",            "Head COM Vertical Offset at BFC",      "cm", -200,  0),
    ("head_d_bfc_cm",             "Head COM Distance at BFC",             "cm",  80, 200),
]

VELOCITY_PARAMS = [
    ("wrist_speed_at_release_m_s","Wrist Speed at Release",               "m/s",  5,  30),
]

FRAME_PARAMS = [
    ("bfc_frame",                 "Back Foot Contact (BFC) Frame",        "fr",   0, 700),
    ("ffc_frame",                 "Front Foot Contact (FFC) Frame",       "fr",   0, 700),
    ("arm_back_frame",            "Arm-Back Event Frame",                 "fr",   0, 700),
    ("release_frame.1",           "Ball Release Frame",                   "fr",   0, 700),
]

ALL_PARAMS = (ELBOW_PARAMS + KNEE_PARAMS + TEMPORAL_PARAMS +
              STRIDE_PARAMS + HEAD_COM_PARAMS + VELOCITY_PARAMS + FRAME_PARAMS)

CATEGORY_MAP = {
    **{p[0]: "Elbow Kinematics"  for p in ELBOW_PARAMS},
    **{p[0]: "Knee Kinematics"   for p in KNEE_PARAMS},
    **{p[0]: "Step & Stride Timing" for p in TEMPORAL_PARAMS},
    **{p[0]: "Delivery Stride"   for p in STRIDE_PARAMS},
    **{p[0]: "Head COM"          for p in HEAD_COM_PARAMS},
    **{p[0]: "Wrist Velocity"    for p in VELOCITY_PARAMS},
    **{p[0]: "Event Frames"      for p in FRAME_PARAMS},
}

# ── Colour palette ───────────────────────────────────────────────────────────
CAT_COLORS = {
    "Elbow Kinematics":    "#2563EB",   # blue
    "Knee Kinematics":     "#0891B2",   # cyan
    "Step & Stride Timing":"#16A34A",   # green
    "Delivery Stride":     "#65A30D",   # lime
    "Head COM":            "#D97706",   # amber
    "Wrist Velocity":      "#9333EA",   # purple
    "Event Frames":        "#DC2626",   # red
}

HEADER_FILL  = "1E3A5F"   # dark navy
SUBHEAD_FILL = "2D6A9F"   # medium blue
CAT_FILLS = {
    "Elbow Kinematics":    "DBEAFE",   # blue tint
    "Knee Kinematics":     "CFFAFE",   # cyan tint
    "Step & Stride Timing":"DCFCE7",   # green tint
    "Delivery Stride":     "ECFCCB",   # lime tint
    "Head COM":            "FEF9C3",   # amber tint
    "Wrist Velocity":      "F3E8FF",   # purple tint
    "Event Frames":        "FFE4E6",   # red tint
}

# ── Utilities ────────────────────────────────────────────────────────────────

def safe_float(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return np.nan
    try:
        return float(v)
    except Exception:
        return np.nan


# ── File paths – edit these to match your local setup ────────────────────────
GT_PATH    = r"C:\Users\nipul\OneDrive\Desktop\tm\dataset_making\New Dataset\ground_truth.csv"
MODEL_PATH = r"C:\Users\nipul\OneDrive\Desktop\tm\dataset_making\New Dataset\masterNewNewFinal.xlsx"
OUTPUT_DIR = r"C:\Users\nipul\OneDrive\Desktop\tm\dataset_making\New Dataset\outputs"
# ─────────────────────────────────────────────────────────────────────────────


def load_data():
    if not os.path.exists(GT_PATH):
        raise FileNotFoundError(f"Ground truth file not found: {GT_PATH}")
    if not os.path.exists(MODEL_PATH):
        raise FileNotFoundError(f"Model file not found: {MODEL_PATH}")

    gt    = pd.read_csv(GT_PATH)
    model = pd.read_excel(MODEL_PATH)
    merged = pd.merge(gt, model, on="trial_id", suffixes=("_gt", "_model"))
    print(f"  GT rows    : {len(gt)}")
    print(f"  Model rows : {len(model)}")
    print(f"  Matched    : {len(merged)}")
    return merged


def icc_2_1(gt_vals, model_vals):
    """Two-way mixed, single measures, absolute agreement (ICC 2,1)."""
    n = len(gt_vals)
    if n < 3:
        return np.nan, np.nan, np.nan
    data = np.column_stack([gt_vals, model_vals])
    grand_mean = data.mean()
    ss_total = np.sum((data - grand_mean) ** 2)
    row_means = data.mean(axis=1)
    ss_rows = 2 * np.sum((row_means - grand_mean) ** 2)
    col_means = data.mean(axis=0)
    ss_cols = n * np.sum((col_means - grand_mean) ** 2)
    ss_error = ss_total - ss_rows - ss_cols
    ms_rows  = ss_rows  / (n - 1)
    ms_error = ss_error / ((n - 1) * 1)  # (n-1)(k-1), k=2
    ms_cols  = ss_cols  / 1
    # absolute agreement formula
    icc = (ms_rows - ms_error) / (ms_rows + ms_error + (2 / n) * (ms_cols - ms_error))
    # 95 % CI via F-distribution
    alpha = 0.05
    F = ms_rows / ms_error
    df1, df2 = n - 1, n - 1
    F_lower = F / stats.f.ppf(1 - alpha / 2, df1, df2)
    F_upper = F * stats.f.ppf(1 - alpha / 2, df2, df1)
    ci_lo = (F_lower - 1) / (F_lower + 1)
    ci_hi = (F_upper - 1) / (F_upper + 1)
    return float(np.clip(icc, -1, 1)), float(ci_lo), float(ci_hi)


def compute_metrics(merged):
    rows = []
    for col, label, unit, lo, hi in ALL_PARAMS:
        gtc   = col + "_gt"    if col + "_gt"    in merged.columns else col
        modc  = col + "_model" if col + "_model" in merged.columns else col
        if gtc not in merged.columns or modc not in merged.columns:
            continue

        gv = merged[gtc].apply(safe_float)
        mv = merged[modc].apply(safe_float)
        mask = gv.notna() & mv.notna()
        if mask.sum() < 2:
            continue

        x, y = gv[mask].values, mv[mask].values
        errors = x - y
        abs_err = np.abs(errors)

        mae  = float(np.mean(abs_err))
        rmse = float(np.sqrt(np.mean(errors ** 2)))
        me   = float(np.mean(errors))    # systematic bias
        std_err = float(np.std(errors))  # random error component

        # Range-normalised RMSE (fair for bounded variables like angles)
        param_range = hi - lo
        n_rmse = rmse / param_range if param_range > 0 else np.nan

        # Pearson r
        if np.std(x) > 0 and np.std(y) > 0:
            r, p_val = stats.pearsonr(x, y)
        else:
            r, p_val = np.nan, np.nan

        # ICC
        icc_val, icc_lo, icc_hi = icc_2_1(x, y)

        # Bland-Altman limits of agreement
        ba_mean   = (x + y) / 2
        ba_diff   = x - y
        ba_bias   = float(np.mean(ba_diff))
        ba_loa_lo = float(ba_bias - 1.96 * np.std(ba_diff))
        ba_loa_hi = float(ba_bias + 1.96 * np.std(ba_diff))

        rows.append({
            "Category":      CATEGORY_MAP.get(col, "Other"),
            "Parameter":     col,
            "Label":         label,
            "Unit":          unit,
            "N":             int(mask.sum()),
            "ME (Bias)":     round(me, 4),
            "MAE":           round(mae, 4),
            "SD Error":      round(std_err, 4),
            "RMSE":          round(rmse, 4),
            "nRMSE (%)":     round(n_rmse * 100, 2) if not np.isnan(n_rmse) else np.nan,
            "Pearson r":     round(r, 4) if not np.isnan(r) else np.nan,
            "p-value":       round(p_val, 4) if not np.isnan(p_val) else np.nan,
            "ICC(2,1)":      round(icc_val, 4) if not np.isnan(icc_val) else np.nan,
            "ICC CI Lo":     round(icc_lo, 4) if not np.isnan(icc_lo) else np.nan,
            "ICC CI Hi":     round(icc_hi, 4) if not np.isnan(icc_hi) else np.nan,
            "BA Bias":       round(ba_bias, 4),
            "BA LoA Lo":     round(ba_loa_lo, 4),
            "BA LoA Hi":     round(ba_loa_hi, 4),
            "_gt_vals":      x,
            "_model_vals":   y,
            "_errors":       errors,
            "_ba_mean":      ba_mean,
            "_ba_diff":      ba_diff,
        })
    return rows


# ── Figure 1 – Scatter plots (identity line) ─────────────────────────────────

def plot_scatter(metrics_rows, out="scatter_plots.png"):
    params = [r for r in metrics_rows if r["Category"] in ("Elbow Kinematics", "Knee Kinematics", "Head COM", "Wrist Velocity", "Delivery Stride")][:12]
    n = len(params)
    cols = 4
    rows_n = (n + cols - 1) // cols
    fig, axes = plt.subplots(rows_n, cols, figsize=(5 * cols, 4.5 * rows_n))
    axes = axes.flatten()

    for i, r in enumerate(params):
        ax = axes[i]
        x, y  = r["_gt_vals"], r["_model_vals"]
        color = CAT_COLORS.get(r["Category"], "#555")

        ax.scatter(x, y, color=color, s=55, alpha=0.75, zorder=3, edgecolors="white", linewidths=0.5)

        lo_v = min(x.min(), y.min()) * 0.93
        hi_v = max(x.max(), y.max()) * 1.07
        ax.plot([lo_v, hi_v], [lo_v, hi_v], "k--", lw=1.2, label="Identity")

        # Regression line
        m_reg, b_reg, *_ = stats.linregress(x, y)
        xfit = np.linspace(lo_v, hi_v, 100)
        ax.plot(xfit, m_reg * xfit + b_reg, color=color, lw=1.4, alpha=0.6, label="Fit")

        ax.set_xlabel("Manual (GT)", fontsize=8)
        ax.set_ylabel("Model", fontsize=8)
        ax.set_title(r["Label"], fontsize=9, fontweight="bold")

        info = (f"MAE={r['MAE']:.3f} {r['Unit']}\n"
                f"RMSE={r['RMSE']:.3f}\n"
                f"r={r['Pearson r']:.3f}")
        ax.text(0.04, 0.97, info, transform=ax.transAxes, va="top", fontsize=7.5,
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.85))
        ax.grid(True, ls="--", alpha=0.35)
        ax.tick_params(labelsize=7)

    for j in range(i + 1, len(axes)):
        axes[j].set_visible(False)

    fig.suptitle("Ground Truth vs Model – Scatter Plots (with Identity & Regression Lines)",
                 fontsize=13, fontweight="bold", y=1.01)
    plt.tight_layout()
    plt.savefig(out, dpi=160, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Figure 2 – Bland-Altman plots ────────────────────────────────────────────

def plot_bland_altman(metrics_rows, out="bland_altman_plots.png"):
    params = [r for r in metrics_rows if r["Category"] in ("Elbow Kinematics", "Knee Kinematics", "Head COM", "Wrist Velocity", "Delivery Stride")][:12]
    n = len(params)
    cols = 4
    rows_n = (n + cols - 1) // cols
    fig, axes = plt.subplots(rows_n, cols, figsize=(5 * cols, 4.5 * rows_n))
    axes = axes.flatten()

    for i, r in enumerate(params):
        ax = axes[i]
        means  = r["_ba_mean"]
        diffs  = r["_ba_diff"]
        bias   = r["BA Bias"]
        loa_lo = r["BA LoA Lo"]
        loa_hi = r["BA LoA Hi"]
        color  = CAT_COLORS.get(r["Category"], "#555")

        ax.scatter(means, diffs, color=color, s=50, alpha=0.75, zorder=3,
                   edgecolors="white", linewidths=0.5)

        ax.axhline(bias,   color="black",  lw=1.4, linestyle="-",  label=f"Bias={bias:.3f}")
        ax.axhline(loa_hi, color="tomato", lw=1.2, linestyle="--", label=f"+1.96σ={loa_hi:.3f}")
        ax.axhline(loa_lo, color="tomato", lw=1.2, linestyle="--", label=f"-1.96σ={loa_lo:.3f}")
        ax.axhline(0, color="grey", lw=0.7, linestyle=":")

        ax.fill_between(ax.get_xlim(), loa_lo, loa_hi, alpha=0.06, color="tomato")
        ax.set_xlabel("Mean of GT & Model", fontsize=8)
        ax.set_ylabel("GT − Model", fontsize=8)
        ax.set_title(r["Label"], fontsize=9, fontweight="bold")
        ax.legend(fontsize=6.5, loc="upper right")
        ax.grid(True, ls="--", alpha=0.35)
        ax.tick_params(labelsize=7)

    for j in range(i + 1, len(axes)):
        axes[j].set_visible(False)

    fig.suptitle("Bland-Altman Agreement Plots  (GT − Model vs Mean)",
                 fontsize=13, fontweight="bold", y=1.01)
    plt.tight_layout()
    plt.savefig(out, dpi=160, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Figure 3 – Error distributions ───────────────────────────────────────────

def plot_error_distributions(metrics_rows, out="error_distribution.png"):
    """Histogram + boxplot for each parameter's signed error."""
    params = [r for r in metrics_rows if r["Category"] in ("Elbow Kinematics", "Knee Kinematics", "Head COM", "Wrist Velocity", "Delivery Stride")][:12]
    n = len(params)
    cols = 4
    rows_n = (n + cols - 1) // cols
    fig, axes = plt.subplots(rows_n, cols, figsize=(5 * cols, 4 * rows_n))
    axes = axes.flatten()

    for i, r in enumerate(params):
        ax = axes[i]
        errs  = r["_errors"]
        color = CAT_COLORS.get(r["Category"], "#555")

        ax.hist(errs, bins=min(12, len(errs)), color=color, alpha=0.7, edgecolor="white")
        ax.axvline(0,         color="black", lw=1.2, linestyle="--", label="Zero error")
        ax.axvline(r["ME (Bias)"], color="tomato", lw=1.2, linestyle="-", label=f"Bias={r['ME (Bias)']:.2f}")

        ax.set_xlabel(f"Error ({r['Unit']})", fontsize=8)
        ax.set_ylabel("Count", fontsize=8)
        ax.set_title(r["Label"], fontsize=9, fontweight="bold")
        ax.legend(fontsize=7)
        ax.grid(True, ls="--", alpha=0.3)
        ax.tick_params(labelsize=7)

    for j in range(i + 1, len(axes)):
        axes[j].set_visible(False)

    fig.suptitle("Error Distributions  (GT − Model)",
                 fontsize=13, fontweight="bold", y=1.01)
    plt.tight_layout()
    plt.savefig(out, dpi=160, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Figure 4 – Summary heatmap ───────────────────────────────────────────────

def plot_summary_heatmap(metrics_rows, out="summary_heatmap.png"):
    labels  = [r["Label"]      for r in metrics_rows]
    mae_v   = [r["MAE"]        for r in metrics_rows]
    nrmse_v = [r["nRMSE (%)"]  for r in metrics_rows]
    icc_v   = [r["ICC(2,1)"]   for r in metrics_rows]
    cats    = [r["Category"]   for r in metrics_rows]

    # nRMSE may be NaN for params with undefined range — replace with 0
    nrmse_v = [v if not (isinstance(v, float) and np.isnan(v)) else 0 for v in nrmse_v]

    # Normalise MAE column to [0,1] for colour mapping
    mae_arr   = np.array(mae_v, dtype=float)
    nrmse_arr = np.array(nrmse_v, dtype=float)
    icc_arr   = np.array(icc_v, dtype=float)

    def norm01(arr):
        r = arr.max() - arr.min()
        return (arr - arr.min()) / r if r > 0 else arr * 0

    fig, axes = plt.subplots(1, 3, figsize=(14, max(6, len(labels) * 0.45)))

    metrics_data = [
        (mae_arr,   "MAE (native units)", "Reds"),
        (nrmse_arr, "nRMSE (%)",           "Oranges"),
        (icc_arr,   "ICC(2,1)",            "RdYlGn"),
    ]

    for ax, (data, title, cmap) in zip(axes, metrics_data):
        img = ax.imshow(data.reshape(-1, 1), aspect="auto", cmap=cmap,
                        vmin=data.min(), vmax=data.max())
        ax.set_yticks(range(len(labels)))
        ax.set_yticklabels(labels, fontsize=8)
        ax.set_xticks([])
        ax.set_title(title, fontsize=10, fontweight="bold", pad=6)

        for j, v in enumerate(data):
            txt = f"{v:.3f}" if not np.isnan(v) else "–"
            ax.text(0, j, txt, ha="center", va="center", fontsize=8,
                    color="white" if (cmap != "RdYlGn" and norm01(data)[j] > 0.65) else "black")

        plt.colorbar(img, ax=ax, fraction=0.04, pad=0.02)

    # Colour y-tick labels by category
    for ax in axes:
        for tick_lbl, cat in zip(ax.get_yticklabels(), cats):
            tick_lbl.set_color(CAT_COLORS.get(cat, "black"))

    fig.suptitle("Summary Heatmap: MAE | nRMSE | ICC",
                 fontsize=13, fontweight="bold", y=1.01)
    plt.tight_layout()
    plt.savefig(out, dpi=160, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Excel export ──────────────────────────────────────────────────────────────

def _style_header(cell, bg=HEADER_FILL):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border    = Border(bottom=thin)


def _col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_metrics_sheet(wb, metrics_rows, raw_df):
    """Sheet 1 – full metrics table."""
    ws = wb.create_sheet("Evaluation Metrics")

    display_cols = [
        "Category", "Label", "Unit", "N",
        "ME (Bias)", "MAE", "SD Error", "RMSE", "nRMSE (%)",
        "Pearson r", "p-value", "ICC(2,1)", "ICC CI Lo", "ICC CI Hi",
        "BA Bias", "BA LoA Lo", "BA LoA Hi",
    ]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_cols))
    title_cell = ws.cell(1, 1, "Cricket Biomechanics – Model Evaluation Metrics")
    title_cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    title_cell.fill      = PatternFill("solid", fgColor=HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for c_idx, col in enumerate(display_cols, 1):
        cell = ws.cell(2, c_idx, col)
        _style_header(cell, SUBHEAD_FILL)

    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    # Data rows
    for r_idx, r in enumerate(metrics_rows, 3):
        cat   = r["Category"]
        fill  = PatternFill("solid", fgColor=CAT_FILLS.get(cat, "FFFFFF"))
        for c_idx, col in enumerate(display_cols, 1):
            val  = r.get(col, "")
            cell = ws.cell(r_idx, c_idx, val)
            cell.fill      = fill
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center" if c_idx > 3 else "left")
            if isinstance(val, float) and not np.isnan(val):
                cell.number_format = "0.0000"

    # Column widths
    widths = [12, 28, 8, 5, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        _col_width(ws, i, w)

    # Conditional formatting on nRMSE column (column 9)
    nrmse_col = get_column_letter(9)
    ws.conditional_formatting.add(
        f"{nrmse_col}3:{nrmse_col}{len(metrics_rows)+2}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B")
    )

    return ws


def _add_raw_sheet(wb, merged_df, metrics_rows):
    """Sheet 2 – trial-level comparison."""
    ws = wb.create_sheet("Trial Comparison")

    param_cols = [(r["Parameter"], r["Label"], r["Unit"]) for r in metrics_rows]
    cols = ["trial_id"]
    headers = ["Trial ID"]
    subheaders = [""]
    for col, label, unit in param_cols:
        gtc  = col + "_gt"
        modc = col + "_model"
        if gtc in merged_df.columns and modc in merged_df.columns:
            cols += [gtc, modc, "__err__" + col]
            headers     += [label, label, label]
            subheaders  += ["GT", "Model", "Error (GT−Model)"]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    tc = ws.cell(1, 1, "Trial-Level Ground Truth vs Model Comparison")
    tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    tc.fill = PatternFill("solid", fgColor=HEADER_FILL)
    tc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Header rows 2 & 3
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(2, c_idx, h)
        _style_header(cell, SUBHEAD_FILL)
    for c_idx, h in enumerate(subheaders, 1):
        cell = ws.cell(3, c_idx, h)
        _style_header(cell, "3B82F6")

    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "B4"

    # Data
    for r_idx, (_, row) in enumerate(merged_df.iterrows(), 4):
        stripe = "F1F5F9" if r_idx % 2 == 0 else "FFFFFF"
        fill   = PatternFill("solid", fgColor=stripe)

        ws.cell(r_idx, 1, row["trial_id"]).font = Font(bold=True, name="Arial", size=9)

        c_idx = 2
        for col, label, unit in param_cols:
            gtc  = col + "_gt"
            modc = col + "_model"
            if gtc not in merged_df.columns or modc not in merged_df.columns:
                continue
            gv = safe_float(row[gtc])
            mv = safe_float(row[modc])
            ev = (gv - mv) if not (np.isnan(gv) or np.isnan(mv)) else np.nan

            for v, c in [(gv, c_idx), (mv, c_idx + 1), (ev, c_idx + 2)]:
                cell = ws.cell(r_idx, c, v if not np.isnan(v) else "")
                cell.fill  = fill
                cell.font  = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="right")
                if isinstance(v, float):
                    cell.number_format = "0.0000"

            # Highlight large errors in red
            err_cell = ws.cell(r_idx, c_idx + 2)
            if not np.isnan(ev) and abs(ev) > 5:
                err_cell.fill = PatternFill("solid", fgColor="FEE2E2")
                err_cell.font = Font(name="Arial", size=9, color="991B1B")

            c_idx += 3

    # Column widths
    ws.column_dimensions["A"].width = 16
    for c_idx in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12

    return ws


def _add_summary_sheet(wb, metrics_rows):
    """Sheet 3 – category summary."""
    ws = wb.create_sheet("Category Summary")

    df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_")}
                       for r in metrics_rows])
    summary = df.groupby("Category").agg(
        Params=("Parameter", "count"),
        MAE_mean=("MAE", "mean"),
        MAE_max=("MAE", "max"),
        RMSE_mean=("RMSE", "mean"),
        nRMSE_mean=("nRMSE (%)", "mean"),
        ICC_mean=("ICC(2,1)", "mean"),
        ICC_min=("ICC(2,1)", "min"),
    ).reset_index()

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    tc = ws.cell(1, 1, "Per-Category Summary Statistics")
    tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    tc.fill = PatternFill("solid", fgColor=HEADER_FILL)
    tc.alignment = Alignment(horizontal="center")

    headers = ["Category", "# Params", "MAE (mean)", "MAE (max)",
               "RMSE (mean)", "nRMSE % (mean)", "ICC mean", "ICC min"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        _style_header(cell, SUBHEAD_FILL)

    for r_idx, row in enumerate(summary.itertuples(index=False), 3):
        cat  = row.Category
        fill = PatternFill("solid", fgColor=CAT_FILLS.get(cat, "FFFFFF"))
        vals = [cat, row.Params, round(row.MAE_mean, 4), round(row.MAE_max, 4),
                round(row.RMSE_mean, 4), round(row.nRMSE_mean, 2),
                round(row.ICC_mean, 4), round(row.ICC_min, 4)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r_idx, c, v)
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")
            if isinstance(v, float):
                cell.number_format = "0.0000"

    widths = [14, 10, 12, 12, 14, 16, 12, 12]
    for i, w in enumerate(widths, 1):
        _col_width(ws, i, w)

    # Interpretation guide
    ws.cell(r_idx + 2, 1, "Interpretation Guide").font = Font(bold=True, name="Arial", size=10)
    guide = [
        ("nRMSE < 5%",    "Excellent agreement"),
        ("nRMSE 5–10%",   "Good agreement"),
        ("nRMSE 10–20%",  "Moderate agreement"),
        ("nRMSE > 20%",   "Poor agreement"),
        ("ICC > 0.90",    "Excellent reliability"),
        ("ICC 0.75–0.90", "Good reliability"),
        ("ICC 0.50–0.75", "Moderate reliability"),
        ("ICC < 0.50",    "Poor reliability"),
    ]
    for j, (threshold, interp) in enumerate(guide, r_idx + 3):
        ws.cell(j, 1, threshold).font = Font(name="Arial", size=9, italic=True)
        ws.cell(j, 2, interp).font    = Font(name="Arial", size=9)

    return ws


def _add_notes_sheet(wb):
    """Sheet 4 – methodology notes."""
    ws = wb.create_sheet("Methodology Notes")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    notes = [
        ("Metric", "Description & Rationale"),
        ("MAE", "Mean Absolute Error in native units. Always interpretable; preferred primary metric."),
        ("ME (Bias)", "Mean signed error (GT − Model). Positive = model underestimates. Reveals systematic offset."),
        ("SD Error", "Standard deviation of signed errors. Represents random/variable component of error."),
        ("RMSE", "Root Mean Square Error; penalises large errors more than MAE. Report alongside MAE."),
        ("nRMSE (%)", ("Range-normalised RMSE = RMSE / (parameter_max − parameter_min) × 100. "
                       "Dimensionless; avoids inflated % errors when GT values are small. "
                       "Recommended for angular data (0–180°). Thresholds: <5% Excellent, 5–10% Good, 10–20% Moderate.")),
        ("Pearson r", "Linear correlation coefficient. High r can coexist with large systematic bias; always pair with BA plot."),
        ("ICC(2,1)", ("Intraclass Correlation Coefficient, two-way mixed, absolute agreement, single measures. "
                      "Accounts for both systematic and random error. "
                      "Thresholds (Koo & Mae 2016): <0.50 Poor, 0.50–0.75 Moderate, 0.75–0.90 Good, >0.90 Excellent.")),
        ("BA Bias", "Bland-Altman mean difference (GT − Model). Quantifies systematic offset across the measurement range."),
        ("BA LoA", ("Limits of Agreement = Bias ± 1.96 × SD. 95% of differences expected within these limits. "
                    "Narrower LoA = better agreement. Essential for clinical/applied significance assessment.")),
        ("Why not %RMSE?", ("Percentage error is misleading for angular data: an error of 12° when GT = 18° gives 67%, "
                            "yet on a 0–180° scale this is only 6.7% of the full range. Use nRMSE instead.")),
        ("Recommended plots", "Bland-Altman (required for method-comparison studies), scatter with identity line, error histograms."),
    ]

    ws.merge_cells("A1:B1")
    title = ws.cell(1, 1, "Evaluation Methodology & Metric Definitions")
    title.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    title.fill = PatternFill("solid", fgColor=HEADER_FILL)
    title.alignment = Alignment(horizontal="center")

    for c, h in enumerate(["Metric", "Description & Rationale"], 1):
        cell = ws.cell(2, c, h)
        _style_header(cell, SUBHEAD_FILL)

    for r_idx, (metric, desc) in enumerate(notes[1:], 3):
        ws.cell(r_idx, 1, metric).font = Font(bold=True, name="Arial", size=9)
        cell = ws.cell(r_idx, 2, desc)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = max(30, len(desc) // 8 * 15)

    return ws


def export_excel(metrics_rows, merged_df, out="evaluation_results.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    _add_metrics_sheet(wb, metrics_rows, merged_df)
    _add_raw_sheet(wb, merged_df, metrics_rows)
    _add_summary_sheet(wb, metrics_rows)
    _add_notes_sheet(wb)

    wb.save(out)
    print(f"  Saved → {out}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("\n── Loading data ──")
    merged = load_data()

    print("\n── Computing metrics ──")
    metrics_rows = compute_metrics(merged)
    print(f"  Parameters evaluated: {len(metrics_rows)}")

    print("\n── Generating figures ──")
    plot_scatter(metrics_rows,             out=os.path.join(OUTPUT_DIR, "scatter_plots.png"))
    plot_bland_altman(metrics_rows,        out=os.path.join(OUTPUT_DIR, "bland_altman_plots.png"))
    plot_error_distributions(metrics_rows, out=os.path.join(OUTPUT_DIR, "error_distribution.png"))
    plot_summary_heatmap(metrics_rows,     out=os.path.join(OUTPUT_DIR, "summary_heatmap.png"))

    print("\n── Exporting Excel ──")
    export_excel(metrics_rows, merged,     out=os.path.join(OUTPUT_DIR, "evaluation_results.xlsx"))

    # Print console table
    print("\n" + "="*88)
    print(f"{'Parameter':<32} {'N':>3} {'MAE':>9} {'RMSE':>9} {'nRMSE%':>8} {'ICC':>7}  Unit")
    print("-"*88)
    for r in metrics_rows:
        nrmse = f"{r['nRMSE (%)']:.2f}" if not (isinstance(r['nRMSE (%)'], float) and np.isnan(r['nRMSE (%)'])) else "  –"
        icc   = f"{r['ICC(2,1)']:.4f}" if not (isinstance(r['ICC(2,1)'], float) and np.isnan(r['ICC(2,1)'])) else "   –"
        print(f"{r['Parameter']:<32} {r['N']:>3} {r['MAE']:>9.4f} {r['RMSE']:>9.4f} {nrmse:>8} {icc:>7}  {r['Unit']}")
    print("="*88)

    print(f"\n── Complete. All outputs saved to: {os.path.abspath(OUTPUT_DIR)}/ ──")


if __name__ == "__main__":
    main()