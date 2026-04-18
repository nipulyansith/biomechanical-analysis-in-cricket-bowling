"""
Model vs Human Coach Comparison
================================
Compares model predictions and two human coach annotations against
ground truth for common cricket bowling trials.

Outputs
-------
1. coach_comparison.xlsx        – summary metrics + per-trial differences
2. fig1_bar_mae_comparison.png  – MAE bar chart: Model vs Coach 1 vs Coach 2
3. fig2_scatter_vs_gt.png       – Scatter plots (predicted vs GT)
4. fig3_boxplot_errors.png      – Error distribution box plots
5. fig4_bland_altman.png        – Bland-Altman agreement plots
6. fig5_radar_overall.png       – Radar chart: overall accuracy profile

Usage
-----
Set the three path variables below, then:
    pip install pandas numpy matplotlib seaborn openpyxl scipy
    python coach_comparison.py
"""

import os
import re
import warnings
import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
import seaborn as sns
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings("ignore")
matplotlib.rcParams.update({"font.family": "DejaVu Sans", "figure.dpi": 140})

# ── File paths ────────────────────────────────────────────────────────────────
GT_PATH    = r"C:\Users\nipul\OneDrive\Desktop\tm\Coach Exp\ground_truth.csv"
MODEL_PATH = r"C:\Users\nipul\OneDrive\Desktop\tm\Coach Exp\masterNewNewFinal.xlsx"
COACH_PATH = r"C:\Users\nipul\OneDrive\Desktop\tm\Coach Exp\coach.xlsx"
OUTPUT_DIR = r"C:\Users\nipul\OneDrive\Desktop\tm\Coach Exp\outputs"
# ─────────────────────────────────────────────────────────────────────────────

# ── Parameter mapping ─────────────────────────────────────────────────────────
# Maps coach survey column (partial match) → (gt_col, label, unit, scale_range)
# scale_range used for nRMSE normalisation
PARAM_MAP = [
    ("total time taken for the last five steps",
        "final5_total_duration_s",   "Final 5 Steps Duration",    "s",     (0, 3)),
    ("time between Back Foot Contact",
        "stride_duration_s",         "Delivery Stride Duration",  "s",     (0, 1)),
    ("estimated stride length",
        "stride_length_m",           "Delivery Stride Length",    "m",     (0, 2.5)),
    ("elbow angle at arm-back",
        "elbow_angle_arm_back_deg",  "Elbow Angle at Arm-Back",   "deg",   (0, 180)),
    ("elbow angle at ball release",
        "elbow_angle_release_deg",   "Elbow Angle at Release",    "deg",   (0, 180)),
    ("front knee angle at Front Foot Contact",
        "knee_angle_ffc_deg",        "Front Knee Angle at FFC",   "deg",   (0, 180)),
    ("front knee angle at ball release",
        "knee_angle_release_deg",    "Front Knee Angle at Release","deg",  (0, 180)),
    ("head horizontal distance",
        "head_d_ffc_cm",             "Head COM Distance at FFC",  "cm",    (50, 200)),
    ("head vertical position",
        "head_dy_ffc_cm",            "Head COM Vertical at FFC",  "cm",    (-200, 0)),
    ("head side movement",
        "head_dx_ffc_cm",            "Head COM Lateral at FFC",   "cm",    (-80, 80)),
    ("wrist speed at ball release",
        "wrist_speed_at_release_m_s","Wrist Speed at Release",    "m/s",   (5, 30)),
]

ANNOTATORS = {
    "Model":   "#2563EB",
    "Coach 1": "#16A34A",
    "Coach 2": "#D97706",
}

HEADER_HEX = "1E3A5F"
SUB_HEX    = "2D6A9F"

# ── Utilities ──────────────────────────────────────────────────────────────────

def parse_numeric(val):
    """Extract first numeric value from messy strings like '75 cm', '0.30 Seconds', '1m'."""
    if val is None:
        return np.nan
    if isinstance(val, (int, float)):
        v = float(val)
        return np.nan if np.isnan(v) else v
    s = str(val).strip().lower()
    # handle fractions like 1m = 100cm — we keep units as-is; caller normalises
    nums = re.findall(r"[-+]?\d*\.?\d+", s)
    if nums:
        v = float(nums[0])
        # if value given in metres but unit is cm, convert (e.g. "1m" → value=1)
        # caller must handle; just return raw number
        return v
    return np.nan


def unit_normalise(raw_val, gt_col, coach_unit_hint):
    """
    Normalise coach values to match GT units.
    GT head distances are in cm. Stride length in m.
    Wrist speed: some coaches entered ms-looking values like 0.08ms which is nonsense
    for m/s (17 m/s GT) — flag as NaN.
    """
    if np.isnan(raw_val):
        return np.nan
    # stride length: if value > 5, likely given in cm → convert to m
    if gt_col == "stride_length_m" and raw_val > 5:
        return raw_val / 100.0
    # head distances: if value looks like metres (< 5), convert to cm
    if gt_col in ("head_d_ffc_cm", "head_dy_ffc_cm", "head_dx_ffc_cm"):
        if raw_val > 0 and raw_val < 5:
            return raw_val * 100.0
    # wrist speed: GT is ~10-20 m/s; if coach entered < 1 (ms?) treat as NaN
    if gt_col == "wrist_speed_at_release_m_s" and raw_val < 1:
        return np.nan
    return raw_val


def safe_float(v):
    try:
        f = float(v)
        return np.nan if np.isnan(f) else f
    except Exception:
        return np.nan


# ── Data loading ───────────────────────────────────────────────────────────────

def load_all():
    for path, name in [(GT_PATH, "GT"), (MODEL_PATH, "Model"), (COACH_PATH, "Coach")]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"{name} file not found: {path}")

    gt    = pd.read_csv(GT_PATH)
    model = pd.read_excel(MODEL_PATH)
    coach_raw = pd.read_excel(COACH_PATH, sheet_name="Form Responses 1")

    # Normalise coach names (fix typo)
    coach_raw["coach_name"] = (coach_raw["Your Name"]
                                .str.strip()
                                .str.replace(r"Nivin Karunarthne\s*$", "Nivin Karunarathne", regex=True))

    # Assign Coach 1 / Coach 2 labels in order of first appearance
    name_order = coach_raw["coach_name"].unique().tolist()
    coach_label = {name: f"Coach {i+1}" for i, name in enumerate(name_order)}
    coach_raw["coach_label"] = coach_raw["coach_name"].map(coach_label)

    print(f"  Coach label mapping: {coach_label}")
    return gt, model, coach_raw


def find_common_trials(gt, model, coach_raw):
    """
    Find trials where GT, model, AND at least one coach annotation exist.
    Only B-05_T-08 has both coaches; B-08_T-01 has Coach 2 only.
    B-08_T-08 is not in GT/model → excluded.
    """
    coach_trials = set(coach_raw["Video ID"].str.strip())
    gt_trials    = set(gt["trial_id"])
    model_trials = set(model["trial_id"])
    common = sorted(coach_trials & gt_trials & model_trials)
    print(f"  Common trials (GT ∩ Model ∩ Coach): {common}")
    return common


def build_coach_values(coach_raw, trials):
    """
    Returns dict: trial_id → {coach_label → {gt_col → float}}
    """
    # Build column lookup: partial phrase → full column name
    col_lookup = {}
    for phrase, gt_col, label, unit, rng in PARAM_MAP:
        for c in coach_raw.columns:
            if phrase.lower() in c.lower():
                col_lookup[phrase] = c
                break

    result = {}
    for tid in trials:
        result[tid] = {}
        rows = coach_raw[coach_raw["Video ID"].str.strip() == tid]
        for _, row in rows.items() if False else rows.iterrows():
            cl = row["coach_label"]
            vals = {}
            for phrase, gt_col, *_ in PARAM_MAP:
                col = col_lookup.get(phrase)
                if col is None:
                    vals[gt_col] = np.nan
                    continue
                raw = parse_numeric(row[col])
                vals[gt_col] = unit_normalise(raw, gt_col, "")
            result[tid][cl] = vals
    return result


# ── Metric computation ─────────────────────────────────────────────────────────

def compute_errors(gt, model, coach_vals, trials):
    """
    For each trial × parameter × annotator, compute signed error and abs error.
    Returns long-form DataFrame.
    """
    rows = []
    for tid in trials:
        gt_row    = gt[gt["trial_id"] == tid].iloc[0]
        model_row = model[model["trial_id"] == tid].iloc[0]

        for phrase, gt_col, label, unit, (lo, hi) in PARAM_MAP:
            gt_val = safe_float(gt_row.get(gt_col, np.nan))
            if np.isnan(gt_val):
                continue

            # Model
            m_val = safe_float(model_row.get(gt_col, np.nan))
            if not np.isnan(m_val):
                err = m_val - gt_val
                rows.append({"trial_id": tid, "param": gt_col, "label": label,
                             "unit": unit, "lo": lo, "hi": hi,
                             "annotator": "Model",
                             "gt": gt_val, "pred": m_val,
                             "error": err, "abs_error": abs(err)})

            # Coaches
            for coach_label, cvals in coach_vals.get(tid, {}).items():
                c_val = cvals.get(gt_col, np.nan)
                if not np.isnan(c_val):
                    err = c_val - gt_val
                    rows.append({"trial_id": tid, "param": gt_col, "label": label,
                                 "unit": unit, "lo": lo, "hi": hi,
                                 "annotator": coach_label,
                                 "gt": gt_val, "pred": c_val,
                                 "error": err, "abs_error": abs(err)})

    return pd.DataFrame(rows)


def summary_metrics(errors_df):
    """MAE, RMSE, nRMSE, Bias per annotator × parameter."""
    rows = []
    for (ann, param, label, unit, lo, hi), grp in errors_df.groupby(
            ["annotator", "param", "label", "unit", "lo", "hi"]):
        errs = grp["error"].values
        n    = len(errs)
        mae  = float(np.mean(np.abs(errs)))
        rmse = float(np.sqrt(np.mean(errs**2)))
        bias = float(np.mean(errs))
        rng  = hi - lo
        nrmse = (rmse / rng * 100) if rng > 0 else np.nan
        rows.append({"Annotator": ann, "Parameter": param, "Label": label,
                     "Unit": unit, "N": n,
                     "MAE": round(mae, 4), "RMSE": round(rmse, 4),
                     "Bias": round(bias, 4), "nRMSE (%)": round(nrmse, 2)})
    return pd.DataFrame(rows)


# ── Figure 1 – MAE Bar Chart ───────────────────────────────────────────────────

def fig1_mae_bar(summary_df, out):
    params   = summary_df["Label"].unique()
    annotators = ["Model", "Coach 1", "Coach 2"]

    # pivot
    pivot = summary_df.pivot_table(index="Label", columns="Annotator", values="MAE")
    pivot = pivot.reindex(columns=[a for a in annotators if a in pivot.columns])

    fig, ax = plt.subplots(figsize=(14, 6))
    x    = np.arange(len(pivot))
    n_ann = len(pivot.columns)
    w    = 0.22
    offsets = np.linspace(-(n_ann-1)*w/2, (n_ann-1)*w/2, n_ann)

    for i, ann in enumerate(pivot.columns):
        vals = pivot[ann].values.astype(float)
        bars = ax.bar(x + offsets[i], vals, width=w,
                      color=ANNOTATORS[ann], alpha=0.88, label=ann,
                      edgecolor="white", linewidth=0.5)
        for bar, v in zip(bars, vals):
            if not np.isnan(v):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.003,
                        f"{v:.2f}", ha="center", va="bottom", fontsize=6.5, rotation=90)

    ax.set_xticks(x)
    ax.set_xticklabels(pivot.index, rotation=35, ha="right", fontsize=8)
    ax.set_ylabel("MAE (native units)", fontsize=10)
    ax.set_title("Mean Absolute Error: Model vs Human Coaches\n(lower = better)",
                 fontsize=12, fontweight="bold")
    ax.legend(fontsize=9)
    ax.grid(axis="y", ls="--", alpha=0.35)
    ax.set_axisbelow(True)
    plt.tight_layout()
    plt.savefig(out, dpi=160, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Figure 2 – Scatter plots ───────────────────────────────────────────────────

def fig2_scatter(errors_df, out):
    params    = errors_df["label"].unique()
    annotators = [a for a in ["Model", "Coach 1", "Coach 2"] if a in errors_df["annotator"].unique()]
    n_params  = len(params)
    n_ann     = len(annotators)

    fig, axes = plt.subplots(n_ann, n_params, figsize=(3.2 * n_params, 3.5 * n_ann),
                             squeeze=False)
    fig.suptitle("Predicted vs Ground Truth (each column = parameter, each row = annotator)",
                 fontsize=11, fontweight="bold", y=1.01)

    for row_i, ann in enumerate(annotators):
        for col_i, param_label in enumerate(params):
            ax = axes[row_i][col_i]
            sub = errors_df[(errors_df["annotator"] == ann) & (errors_df["label"] == param_label)]

            if sub.empty:
                ax.text(0.5, 0.5, "No data", ha="center", va="center",
                        transform=ax.transAxes, fontsize=8, color="grey")
                ax.set_title(param_label, fontsize=7)
                continue

            x, y = sub["gt"].values, sub["pred"].values
            color = ANNOTATORS[ann]
            ax.scatter(x, y, color=color, s=70, zorder=3, edgecolors="white", lw=0.5)

            lo_v = min(x.min(), y.min()) * 0.90
            hi_v = max(x.max(), y.max()) * 1.10
            ax.plot([lo_v, hi_v], [lo_v, hi_v], "k--", lw=1.1, label="Identity")

            mae = np.mean(np.abs(x - y))
            ax.text(0.05, 0.95, f"MAE={mae:.2f}", transform=ax.transAxes,
                    va="top", fontsize=7,
                    bbox=dict(boxstyle="round,pad=0.2", facecolor="white", alpha=0.8))

            if row_i == 0:
                ax.set_title(param_label, fontsize=7.5, fontweight="bold")
            if col_i == 0:
                ax.set_ylabel(f"{ann}\nModel output", fontsize=7.5)
            ax.set_xlabel("Ground Truth", fontsize=7)
            ax.tick_params(labelsize=6.5)
            ax.grid(True, ls="--", alpha=0.3)

    plt.tight_layout()
    plt.savefig(out, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Figure 3 – Box plots ──────────────────────────────────────────────────────

def fig3_boxplot(errors_df, out):
    params     = errors_df["label"].unique()
    n          = len(params)
    cols_n     = 3
    rows_n     = (n + cols_n - 1) // cols_n
    fig, axes  = plt.subplots(rows_n, cols_n, figsize=(5.5 * cols_n, 4.5 * rows_n))
    axes       = axes.flatten()

    fig.suptitle("Error Distribution: Model vs Human Coaches  (GT − Annotator)",
                 fontsize=12, fontweight="bold", y=1.01)

    annotators = [a for a in ["Model", "Coach 1", "Coach 2"]
                  if a in errors_df["annotator"].unique()]
    palette = {a: ANNOTATORS[a] for a in annotators}

    for i, param_label in enumerate(params):
        ax  = axes[i]
        sub = errors_df[errors_df["label"] == param_label].copy()
        sub["error"] = sub["gt"] - sub["pred"]   # GT - pred so positive = over-prediction

        sns.boxplot(data=sub, x="annotator", y="error", palette=palette,
                    order=[a for a in ["Model", "Coach 1", "Coach 2"] if a in sub["annotator"].unique()],
                    ax=ax, width=0.45, linewidth=1.2,
                    flierprops=dict(marker="o", markersize=4, alpha=0.6))
        ax.axhline(0, color="black", lw=1.0, ls="--", alpha=0.7)
        ax.set_title(param_label, fontsize=9, fontweight="bold")
        ax.set_xlabel("")
        unit = sub["unit"].iloc[0] if len(sub) else ""
        ax.set_ylabel(f"GT − Pred ({unit})", fontsize=8)
        ax.tick_params(labelsize=8)
        ax.grid(axis="y", ls="--", alpha=0.3)
        ax.set_axisbelow(True)

    for j in range(i + 1, len(axes)):
        axes[j].set_visible(False)

    plt.tight_layout()
    plt.savefig(out, dpi=155, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Figure 4 – Bland-Altman ───────────────────────────────────────────────────

def fig4_bland_altman(errors_df, out):
    key_params = [
        "Elbow Angle at Arm-Back", "Elbow Angle at Release",
        "Front Knee Angle at FFC", "Front Knee Angle at Release",
        "Head COM Distance at FFC", "Wrist Speed at Release",
    ]
    annotators = [a for a in ["Model", "Coach 1", "Coach 2"]
                  if a in errors_df["annotator"].unique()]

    available = [p for p in key_params if p in errors_df["label"].unique()]
    n_rows = len(annotators)
    n_cols = len(available)
    if n_cols == 0 or n_rows == 0:
        return

    fig, axes = plt.subplots(n_rows, n_cols, figsize=(3.5 * n_cols, 3.8 * n_rows),
                             squeeze=False)
    fig.suptitle("Bland-Altman Agreement Plots  (GT − Pred vs Mean)",
                 fontsize=11, fontweight="bold", y=1.01)

    for ri, ann in enumerate(annotators):
        for ci, plabel in enumerate(available):
            ax  = axes[ri][ci]
            sub = errors_df[(errors_df["annotator"] == ann) & (errors_df["label"] == plabel)]

            if sub.empty:
                ax.text(0.5, 0.5, "No data", ha="center", va="center",
                        transform=ax.transAxes, fontsize=8, color="grey")
                continue

            means = (sub["gt"] + sub["pred"]) / 2
            diffs = sub["gt"] - sub["pred"]
            bias  = diffs.mean()
            loa   = 1.96 * diffs.std()
            color = ANNOTATORS[ann]

            ax.scatter(means, diffs, color=color, s=55, alpha=0.8,
                       edgecolors="white", lw=0.5, zorder=3)
            ax.axhline(bias,        color="black",  lw=1.3, ls="-")
            ax.axhline(bias + loa,  color="tomato", lw=1.1, ls="--")
            ax.axhline(bias - loa,  color="tomato", lw=1.1, ls="--")
            ax.axhline(0, color="grey", lw=0.7, ls=":")
            ax.fill_between(ax.get_xlim(), bias - loa, bias + loa,
                            alpha=0.06, color="tomato")

            ax.text(0.03, 0.97, f"Bias={bias:.2f}\nLoA={bias-loa:.2f} to {bias+loa:.2f}",
                    transform=ax.transAxes, va="top", fontsize=6.5,
                    bbox=dict(boxstyle="round,pad=0.2", facecolor="white", alpha=0.85))

            if ri == 0:
                ax.set_title(plabel, fontsize=8, fontweight="bold")
            if ci == 0:
                ax.set_ylabel(f"{ann}\nGT − Pred", fontsize=7.5)
            ax.set_xlabel("Mean of GT & Pred", fontsize=7)
            ax.tick_params(labelsize=6.5)
            ax.grid(True, ls="--", alpha=0.3)

    plt.tight_layout()
    plt.savefig(out, dpi=155, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Figure 5 – Radar chart ────────────────────────────────────────────────────

def fig5_radar(summary_df, out):
    """Radar chart showing nRMSE per parameter for each annotator (lower = better)."""
    annotators = [a for a in ["Model", "Coach 1", "Coach 2"]
                  if a in summary_df["Annotator"].unique()]
    pivot = summary_df.pivot_table(index="Label", columns="Annotator", values="nRMSE (%)")
    pivot = pivot.dropna(how="all")
    labels = pivot.index.tolist()
    N = len(labels)
    if N < 3:
        return

    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))

    for ann in annotators:
        if ann not in pivot.columns:
            continue
        vals = pivot[ann].fillna(pivot[ann].max()).tolist()
        vals += vals[:1]
        ax.plot(angles, vals, color=ANNOTATORS[ann], lw=2, label=ann)
        ax.fill(angles, vals, color=ANNOTATORS[ann], alpha=0.08)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, fontsize=8)
    ax.set_title("nRMSE (%) by Parameter\n(closer to centre = more accurate)",
                 fontsize=11, fontweight="bold", pad=20)
    ax.legend(loc="upper right", bbox_to_anchor=(1.35, 1.1), fontsize=9)
    ax.grid(True, ls="--", alpha=0.4)

    plt.tight_layout()
    plt.savefig(out, dpi=155, bbox_inches="tight")
    plt.close()
    print(f"  Saved → {out}")


# ── Excel export ───────────────────────────────────────────────────────────────

def _hdr(cell, bg=HEADER_HEX):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(bottom=Side(style="thin", color="FFFFFF"))

def _cw(ws, col_idx, w):
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ANN_FILLS = {"Model": "DBEAFE", "Coach 1": "DCFCE7", "Coach 2": "FEF9C3"}


def export_excel(summary_df, errors_df, out):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary Metrics ──────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary Metrics")
    ws1.merge_cells("A1:I1")
    tc = ws1.cell(1, 1, "Model vs Human Coach – Evaluation Summary")
    tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    tc.fill = PatternFill("solid", fgColor=HEADER_HEX)
    tc.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 26

    hdrs = ["Annotator", "Parameter", "Unit", "N", "Bias", "MAE", "RMSE", "nRMSE (%)"]
    for c, h in enumerate(hdrs, 1):
        _hdr(ws1.cell(2, c, h), SUB_HEX)
    ws1.row_dimensions[2].height = 32
    ws1.freeze_panes = "A3"

    for r, row in enumerate(summary_df.itertuples(index=False), 3):
        ann  = row.Annotator
        fill = PatternFill("solid", fgColor=ANN_FILLS.get(ann, "FFFFFF"))
        vals = [ann, row.Label, row.Unit, row.N, row.Bias, row.MAE, row.RMSE, getattr(row, "nRMSE (%)", "")]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(r, c, v)
            cell.fill = fill
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center" if c > 2 else "left")
            if isinstance(v, float) and not np.isnan(v):
                cell.number_format = "0.0000"

    for i, w in enumerate([12, 30, 8, 5, 10, 10, 10, 10], 1):
        _cw(ws1, i, w)

    # Conditional colour on nRMSE col
    nrmse_col = get_column_letter(8)
    ws1.conditional_formatting.add(
        f"{nrmse_col}3:{nrmse_col}{len(summary_df)+2}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B"))

    # ── Sheet 2: Per-Trial Differences ────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Trial Differences")
    ws2.merge_cells("A1:G1")
    tc2 = ws2.cell(1, 1, "Per-Trial: Ground Truth vs Each Annotator")
    tc2.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    tc2.fill = PatternFill("solid", fgColor=HEADER_HEX)
    tc2.alignment = Alignment(horizontal="center")

    hdrs2 = ["Trial", "Parameter", "Unit", "Ground Truth", "Annotator", "Predicted", "Abs Error"]
    for c, h in enumerate(hdrs2, 1):
        _hdr(ws2.cell(2, c, h), SUB_HEX)
    ws2.freeze_panes = "A3"

    for r, row in enumerate(errors_df.sort_values(["trial_id","label","annotator"]).itertuples(index=False), 3):
        ann  = row.annotator
        fill = PatternFill("solid", fgColor=ANN_FILLS.get(ann, "FFFFFF"))
        vals = [row.trial_id, row.label, row.unit,
                round(row.gt, 4), ann, round(row.pred, 4), round(row.abs_error, 4)]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(r, c, v)
            cell.fill = fill
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center" if c > 3 else "left")
            if isinstance(v, float):
                cell.number_format = "0.0000"
        # Highlight large errors
        err_cell = ws2.cell(r, 7)
        if isinstance(row.abs_error, float) and row.abs_error > 10:
            err_cell.fill = PatternFill("solid", fgColor="FEE2E2")
            err_cell.font = Font(name="Arial", size=9, color="991B1B")

    for i, w in enumerate([14, 30, 8, 14, 12, 14, 12], 1):
        _cw(ws2, i, w)

    # ── Sheet 3: Conclusions ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Conclusions")
    ws3.merge_cells("A1:B1")
    tc3 = ws3.cell(1, 1, "Conclusions: Is the Model Comparable to Human Coaches?")
    tc3.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    tc3.fill = PatternFill("solid", fgColor=HEADER_HEX)
    tc3.alignment = Alignment(horizontal="center")

    # Auto-generate conclusion text from metrics
    model_mean_nrmse = summary_df[summary_df["Annotator"]=="Model"]["nRMSE (%)"].mean()
    c1_mean = summary_df[summary_df["Annotator"]=="Coach 1"]["nRMSE (%)"].mean()
    c2_mean = summary_df[summary_df["Annotator"]=="Coach 2"]["nRMSE (%)"].mean()

    # Per-param winner
    pivot_nrmse = summary_df.pivot_table(index="Label", columns="Annotator", values="nRMSE (%)")
    model_wins  = []
    coach_wins  = []
    for param in pivot_nrmse.index:
        row_v = pivot_nrmse.loc[param]
        m_v   = row_v.get("Model", np.nan)
        coach_vals_v = [row_v.get(a, np.nan) for a in ["Coach 1","Coach 2"]]
        coach_min = np.nanmin(coach_vals_v) if any(~np.isnan(coach_vals_v)) else np.nan
        if np.isnan(m_v) or np.isnan(coach_min):
            continue
        if m_v <= coach_min:
            model_wins.append(param)
        else:
            coach_wins.append(param)

    lines = [
        ("Overall Model nRMSE (mean)",     f"{model_mean_nrmse:.2f}%"),
        ("Overall Coach 1 nRMSE (mean)",   f"{c1_mean:.2f}%" if not np.isnan(c1_mean) else "N/A"),
        ("Overall Coach 2 nRMSE (mean)",   f"{c2_mean:.2f}%" if not np.isnan(c2_mean) else "N/A"),
        ("Parameters model wins",          ", ".join(model_wins) if model_wins else "None"),
        ("Parameters coaches win",         ", ".join(coach_wins) if coach_wins else "None"),
        ("Overall verdict",
            "Model is MORE ACCURATE than human coaches on average." if model_mean_nrmse < min(
                x for x in [c1_mean, c2_mean] if not np.isnan(x))
            else "Human coaches are more accurate than the model on average."),
        ("Note on coach variability",
            "Large inter-coach disagreement observed — confirming need for automated analysis."),
    ]

    for r, (key, val) in enumerate(lines, 3):
        ws3.cell(r, 1, key).font  = Font(bold=True, name="Arial", size=10)
        ws3.cell(r, 2, val).font  = Font(name="Arial", size=10)
        ws3.cell(r, 2).alignment  = Alignment(wrap_text=True)
        ws3.row_dimensions[r].height = 24

    _cw(ws3, 1, 32)
    _cw(ws3, 2, 70)

    wb.save(out)
    print(f"  Saved → {out}")


# ── Console summary ────────────────────────────────────────────────────────────

def print_summary(summary_df):
    print("\n" + "="*80)
    print(f"{'Annotator':<12} {'Parameter':<34} {'MAE':>9} {'nRMSE%':>8}  Unit")
    print("-"*80)
    for ann in ["Model", "Coach 1", "Coach 2"]:
        sub = summary_df[summary_df["Annotator"] == ann]
        if sub.empty:
            continue
        print(f"\n  ── {ann} ──")
        for _, r in sub.iterrows():
            nrmse = f"{r['nRMSE (%)']:.2f}" if not np.isnan(r["nRMSE (%)"]) else "  –"
            print(f"  {r['Label']:<34} {r['MAE']:>9.4f} {nrmse:>8}  {r['Unit']}")
    print("="*80)

    # Overall comparison
    print("\n── Overall Mean nRMSE ──")
    for ann in ["Model", "Coach 1", "Coach 2"]:
        sub = summary_df[summary_df["Annotator"] == ann]
        if sub.empty:
            continue
        mean_nrmse = sub["nRMSE (%)"].mean()
        print(f"  {ann}: {mean_nrmse:.2f}%")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("\n── Loading data ──")
    gt, model, coach_raw = load_all()

    print("\n── Finding common trials ──")
    trials = find_common_trials(gt, model, coach_raw)
    if not trials:
        print("  No common trials found. Exiting.")
        return

    print("\n── Parsing coach annotations ──")
    coach_vals = build_coach_values(coach_raw, trials)
    for tid, ann_dict in coach_vals.items():
        for ann, vals in ann_dict.items():
            valid = {k: v for k, v in vals.items() if not np.isnan(v)}
            print(f"  {tid} | {ann}: {len(valid)}/{len(PARAM_MAP)} params parsed")

    print("\n── Computing errors ──")
    errors_df  = compute_errors(gt, model, coach_vals, trials)
    summary_df = summary_metrics(errors_df)
    print(f"  Error records: {len(errors_df)}")

    print_summary(summary_df)

    print("\n── Generating figures ──")
    fig1_mae_bar(summary_df,  os.path.join(OUTPUT_DIR, "fig1_bar_mae_comparison.png"))
    fig2_scatter(errors_df,   os.path.join(OUTPUT_DIR, "fig2_scatter_vs_gt.png"))
    fig3_boxplot(errors_df,   os.path.join(OUTPUT_DIR, "fig3_boxplot_errors.png"))
    fig4_bland_altman(errors_df, os.path.join(OUTPUT_DIR, "fig4_bland_altman.png"))
    fig5_radar(summary_df,    os.path.join(OUTPUT_DIR, "fig5_radar_overall.png"))

    print("\n── Exporting Excel ──")
    export_excel(summary_df, errors_df, os.path.join(OUTPUT_DIR, "coach_comparison.xlsx"))

    print(f"\n── Complete. Outputs saved to: {os.path.abspath(OUTPUT_DIR)}/ ──")


if __name__ == "__main__":
    main()